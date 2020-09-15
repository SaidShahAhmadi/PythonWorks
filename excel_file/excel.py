import openpyxl

wb = openpyxl.load_workbook("book1.xlsx")
# show the current book sheet how many sheet we in the book
print(wb.sheetnames)

# reading data from sheet
sheet = wb["Sheet1"]
cell = sheet["a1"]

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)


# adding new data to sheet
# sheet.appent([1, 2, 3])
# wb.save("updatedbook.xlsx")

# print(cell.row)
# print(cell.column)
